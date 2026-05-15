import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

def parsear_horario_visual(file):
    # Cargar Excel leyendo solo los valores
    wb = openpyxl.load_workbook(file, data_only=True)
    
    nombre_hoja = "HorarioSem"
    if nombre_hoja not in wb.sheetnames:
        st.error(f"No se encontró la hoja '{nombre_hoja}'.")
        return None
    
    ws = wb[nombre_hoja]
    datos_lista = []
    # Columnas: B=2, C=3, D=4, E=5, F=6
    dias = {2: "Lunes", 3: "Martes", 4: "Miércoles", 5: "Jueves", 6: "Viernes"}
    
    bloques = []
    for row in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=1).value).upper()
        if "SEMESTRE" in cell_value and "HORARIOS 202602" not in cell_value:
            bloques.append({"nombre": cell_value, "inicio": row + 2, "fin": row + 14})

    for b in bloques:
        for fila in range(b["inicio"], b["fin"] + 1):
            hora = ws.cell(row=fila, column=1).value
            if not hora: continue
            
            for col_idx, dia_nombre in dias.items():
                contenido = ws.cell(row=fila, column=col_idx).value
                if contenido:
                    lineas = [l.strip() for l in str(contenido).split('\n')]
                    while len(lineas) < 4: lineas.append("PENDIENTE")
                    
                    datos_lista.append({
                        "Semestre": b["nombre"], "Dia": dia_nombre, "Hora": hora,
                        "Codigo": lineas[0], "Asignatura": lineas[1],
                        "Profesor": lineas[2], "Salon": lineas[3]
                    })
                    
    return pd.DataFrame(datos_lista)

# --- INTERFAZ STREAMLIT ---
st.set_page_config(page_title="Gestor de Horarios Universitarios", layout="wide")
st.title("📅 Generador de Horarios por Profesor")

archivo_subido = st.file_uploader("Elige tu archivo Excel (.xlsx)", type=["xlsx"])

if archivo_subido:
    df = parsear_horario_visual(archivo_subido)
    
    if df is not None and not df.empty:
        etiquetas_ignorar = ["PENDIENTE", "POR ASIGNAR", "-", "TBD", ""]
        df_reales = df[~df['Profesor'].isin(etiquetas_ignorar)]
        
        # Validación de Cruces: Solo si el código de asignatura es diferente
        cruces = df_reales[df_reales.groupby(['Dia', 'Hora', 'Profesor'])['Codigo'].transform('nunique') > 1]

        if not cruces.empty:
            st.error("⚠️ Se detectaron conflictos de asignación (Diferentes materias a la misma hora).")
            reporte_df = cruces[['Profesor', 'Dia', 'Hora', 'Asignatura', 'Semestre']].sort_values(by=['Profesor', 'Dia', 'Hora'])
            st.dataframe(reporte_df)
            
            st.download_button(
                label="📥 Descargar Reporte de Errores", 
                data="CONFLICTOS:\n" + reporte_df.to_string(index=False), 
                file_name="Reporte_Errores.txt"
            )
        
        elif df_reales.empty:
            st.warning("No se encontraron profesores reales para procesar (solo hay PENDIENTES o TBD).")
            
        else:
            st.success("✅ ¡Base de datos validada! Generando reporte consolidado...")
            
            output = BytesIO()
            horas_orden = df['Hora'].unique()
            
            try:
                # Lista para recolectar todos los bloques de profesores
                lista_bloques = []

                for profe in df_reales['Profesor'].unique():
                    df_profe = df_reales[df_reales['Profesor'] == profe]
                    
                    # 1. Crear el pivote (horario) del profesor
                    cal = df_profe.pivot_table(
                        index='Hora', 
                        columns='Dia', 
                        values='Asignatura', 
                        aggfunc=lambda x: " / ".join(set(x))
                    )
                    
                    # 2. Reindexar para tener la grilla completa (Lunes a Viernes)
                    cal = cal.reindex(index=horas_orden, columns=["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]).fillna("-")
                    
                    # --- CONSTRUCCIÓN DEL CUADRO INDIVIDUAL ---
                    
                    # Fila A: Título del Docente
                    # Creamos un DF de una fila que ocupa todas las columnas
                    titulo = pd.DataFrame([["" for _ in range(6)]], columns=['H'] + list(cal.columns))
                    titulo.iloc[0, 0] = f"DOCENTE: {profe}"
                    
                    # Fila B: Encabezados de los días (esto es lo que pediste que se repita)
                    encabezados = pd.DataFrame([["HORA", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]], 
                                               columns=['H'] + list(cal.columns))
                    
                    # Fila C: Los datos del horario
                    # Reseteamos el índice de 'cal' para que la columna 'Hora' sea datos reales
                    datos_horario = cal.reset_index()
                    datos_horario.columns = ['H'] + list(cal.columns) # Normalizamos nombres de columnas para el concat
                    
                    # Fila D: Espacio en blanco para separar del siguiente profe
                    espacio = pd.DataFrame([["" for _ in range(6)]], columns=['H'] + list(cal.columns))

                    # Unimos las piezas del profesor actual
                    lista_bloques.append(titulo)
                    lista_bloques.append(encabezados)
                    lista_bloques.append(datos_horario)
                    lista_bloques.append(espacio)

                # Concatenar todos los profesores en un solo gran DataFrame
                if lista_bloques:
                    df_final = pd.concat(lista_bloques, ignore_index=True)

                    # Escribir a Excel
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # index=False y header=False porque ya nosotros creamos los títulos manualmente
                        df_final.to_excel(writer, sheet_name="Horarios_Consolidados", index=False, header=False)

                    st.download_button(
                        label="📥 Descargar Reporte Consolidado (Excel)",
                        data=output.getvalue(),
                        file_name="Reporte_Horarios_Docentes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Error al generar el Excel: {e}")
    else:
        st.info("El archivo parece estar vacío o no contiene la palabra 'SEMESTRE' en la columna A.")
